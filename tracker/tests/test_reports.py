from django.urls import reverse
from django.utils import timezone

from tracker.models import UsageLog

from .base import BaseTestCase


class ReportsFilterTests(BaseTestCase):
    def setUp(self):
        super().setUp()
        UsageLog.objects.create(employee=self.emp1, item=self.goggles, logged_by=self.supervisor)
        UsageLog.objects.create(employee=self.emp1, item=self.gloves, logged_by=self.supervisor)
        UsageLog.objects.create(employee=self.emp2, item=self.earplugs, logged_by=self.supervisor)

    def _tbody(self, response):
        content = response.content.decode()
        return content.split("<tbody>")[1].split("</tbody>")[0]

    def test_usage_log_shows_all_by_default(self):
        self.login_admin()
        response = self.client.get(reverse("reports"), {"type": "usage"})
        tbody = self._tbody(response)
        self.assertIn("EMP-001", tbody)
        self.assertIn("EMP-002", tbody)

    def test_filter_by_employee(self):
        self.login_admin()
        response = self.client.get(reverse("reports"), {"type": "usage", "employee": self.emp1.pk})
        tbody = self._tbody(response)
        self.assertIn("EMP-001", tbody)
        self.assertNotIn("EMP-002", tbody)

    def test_filter_by_item(self):
        self.login_admin()
        response = self.client.get(reverse("reports"), {"type": "usage", "item": self.goggles.pk})
        tbody = self._tbody(response)
        self.assertIn("Safety Goggles", tbody)
        self.assertNotIn("Ear Plugs", tbody)

    def test_future_date_from_returns_no_results(self):
        self.login_admin()
        response = self.client.get(reverse("reports"), {"type": "usage", "date_from": "2099-01-01"})
        self.assertContains(response, "No usage logs match")

    def test_supervisor_can_view_reports(self):
        self.login_supervisor()
        response = self.client.get(reverse("reports"))
        self.assertEqual(response.status_code, 200)


class ReportsPaginationTests(BaseTestCase):
    def setUp(self):
        super().setUp()
        for _ in range(30):
            UsageLog.objects.create(employee=self.emp1, item=self.goggles, logged_by=self.supervisor)

    def test_page_one_has_25_rows(self):
        self.login_admin()
        response = self.client.get(reverse("reports"), {"type": "usage"})
        content = response.content.decode()
        tbody = content.split("<tbody>")[1].split("</tbody>")[0]
        self.assertEqual(tbody.count("<tr>"), 25)

    def test_page_two_has_remaining_rows(self):
        self.login_admin()
        response = self.client.get(reverse("reports"), {"type": "usage", "page": 2})
        content = response.content.decode()
        tbody = content.split("<tbody>")[1].split("</tbody>")[0]
        self.assertEqual(tbody.count("<tr>"), 5)


class ByEmployeeReportTests(BaseTestCase):
    def setUp(self):
        super().setUp()
        now = timezone.now()
        UsageLog.objects.create(employee=self.emp1, item=self.goggles, logged_by=self.supervisor, logged_at=now)
        UsageLog.objects.create(employee=self.emp1, item=self.gloves, logged_by=self.supervisor, logged_at=now)
        UsageLog.objects.create(employee=self.emp2, item=self.earplugs, logged_by=self.supervisor, logged_at=now)

    def test_counts_match_actual_logs(self):
        self.login_admin()
        response = self.client.get(reverse("reports"), {"type": "employee"})
        content = response.content.decode()
        tbody = content.split("<tbody>")[1].split("</tbody>")[0]
        # emp1 has 2 logs today -> should show 2, 2, 2 for today/week/month
        import re
        rows = re.findall(r"<tr>.*?</tr>", tbody, re.S)
        emp1_row = next(r for r in rows if "EMP-001" in r)
        nums = re.findall(r"<td>(\d+)</td>", emp1_row)
        self.assertEqual(nums, ["2", "2", "2"])

    def test_search_filters_by_name(self):
        self.login_admin()
        response = self.client.get(reverse("reports"), {"type": "employee", "q": "Ramesh"})
        content = response.content.decode()
        tbody = content.split("<tbody>")[1].split("</tbody>")[0]
        self.assertIn("Ramesh", tbody)
        self.assertNotIn("Suresh", tbody)

    def test_detail_page_shows_only_date_and_item(self):
        self.login_admin()
        response = self.client.get(reverse("reports_employee_detail", args=[self.emp1.pk]))
        self.assertContains(response, "<th>Date</th>")
        self.assertContains(response, "<th>Item</th>")
        self.assertNotContains(response, "<th>Quantity</th>")

    def test_detail_page_date_filter(self):
        self.login_admin()
        response = self.client.get(
            reverse("reports_employee_detail", args=[self.emp1.pk]),
            {"date_from": "2099-01-01"},
        )
        self.assertContains(response, "No items logged")


class ReportsExportTests(BaseTestCase):
    def setUp(self):
        super().setUp()
        UsageLog.objects.create(employee=self.emp1, item=self.goggles, logged_by=self.supervisor)

    def test_csv_export_usage(self):
        self.login_admin()
        response = self.client.get(reverse("reports_export_csv"), {"type": "usage"})
        self.assertEqual(response["Content-Type"], "text/csv")
        content = response.content.decode()
        self.assertIn("Ramesh Kumar", content)

    def test_csv_export_employee_summary(self):
        self.login_admin()
        response = self.client.get(reverse("reports_export_csv"), {"type": "employee"})
        content = response.content.decode()
        self.assertIn("Employee,Code,Status,Today,This Week,This Month", content)

    def test_xlsx_export_usage_is_valid_workbook(self):
        self.login_admin()
        response = self.client.get(reverse("reports_export_xlsx"), {"type": "usage"})
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws["A1"].value, "Date/Time")
        self.assertTrue(ws["A1"].font.bold)

    def test_xlsx_export_employee_summary(self):
        self.login_admin()
        response = self.client.get(reverse("reports_export_xlsx"), {"type": "employee"})
        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws.title, "Employee Summary")
